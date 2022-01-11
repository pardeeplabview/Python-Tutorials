import openpyxl as op
c = op.load_workbook("D:\\pass.xlsx")
print(c.sheetnames)
s = c['Sheet1']
Password = s['A2'].value

print(Password)
if Password == "=CHAR(RANDBETWEEN(65,90))":
    print("Welcome")
else:
    print("Please Activate this software")
#x = s.cell(row=2,column=1).value
#print(s.max_row)
#print(s.max_column)

#d = s.max_column
#for i in range(1,d+1):
#    print(s.cell(row = 2,column = i).value)

#d1 = s.max_row
#for i in range(1,d1+1):
#    print(s.cell(row = i,column = i).value)
