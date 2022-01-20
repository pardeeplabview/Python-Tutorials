# openpyxl
import openpyxl as op
c = op.load_workbook('D:\\Test\\Book1.xlsx')
print(c.sheetnames)
s = c['Sheet1']
#x = s.cell(row=2, column=2).value
x = s['B2'].value
#print(x)

d1 = s.max_row
print(d1)
for i in range(1,d1+1):
    print(s.cell(row=3,column=i).value)